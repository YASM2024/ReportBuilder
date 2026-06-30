"""設定ファイルの読み込みと検証。"""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Any

import yaml
from openpyxl.utils import column_index_from_string, get_column_letter

from .paths import get_app_root, resolve_app_path


@dataclass
class ColumnMapping:
    csv_column: str | int
    excel_column: str
    number_format: str | None = None
    align: str | None = None  # left | center | right


COLUMN_ALIGNS = ("left", "center", "right")


@dataclass
class SortRule:
    csv_column: str | int
    order: str  # "asc" or "desc"


@dataclass
class HorizontalBorder:
    row: int | str  # "last_data_row" でデータ最終行
    from_col: str
    to_col: str
    style: str = "thin"


@dataclass
class VerticalBorder:
    col: str
    from_row: int
    to_row: int | str  # "last_data_row" でデータ最終行まで延長
    style: str = "thin"


BORDER_MODES = ("outline", "horizontal", "vertical", "grid")


@dataclass
class BorderRegion:
    from_row: int
    to_row: int | str  # "last_data_row" でデータ最終行まで延長
    from_col: str
    to_col: str
    mode: str  # outline | horizontal | vertical | grid
    style: str = "thin"
    internal: bool = True
    row_step: int = 1
    col_step: int = 1


@dataclass
class OutputSettings:
    base_name: str
    include_timestamp: bool = True
    timestamp_format: str = "%Y%m%d_%H%M%S"


@dataclass
class CsvSettings:
    encoding: str = "utf-8-sig"
    delimiter: str = ","
    pattern: str = "*.csv"


@dataclass
class ReportConfig:
    name: str
    config_path: Path
    base_dir: Path
    csv_input_dir: Path
    excel_output_dir: Path
    template_path: Path
    output: OutputSettings
    csv: CsvSettings
    sheet_name: str | None
    start_row: int
    columns: list[ColumnMapping]
    sort_rules: list[SortRule]
    horizontal_borders: list[HorizontalBorder]
    vertical_borders: list[VerticalBorder]
    border_regions: list[BorderRegion]


def _require(mapping: dict[str, Any], key: str, context: str) -> Any:
    if key not in mapping:
        raise ValueError(f"{context}: '{key}' が必要です")
    return mapping[key]


def _parse_columns(raw: list[dict[str, Any]]) -> list[ColumnMapping]:
    columns: list[ColumnMapping] = []
    for index, item in enumerate(raw):
        csv_col = item.get("csv")
        if csv_col is None:
            raise ValueError(f"columns[{index}]: 'csv' が必要です")
        excel_col = _require(item, "excel", f"columns[{index}]")
        column_index_from_string(str(excel_col).upper())
        align = item.get("align")
        if align is not None:
            align = str(align).lower()
            if align not in COLUMN_ALIGNS:
                allowed = ", ".join(COLUMN_ALIGNS)
                raise ValueError(
                    f"columns[{index}].align は {allowed} のいずれかを指定してください"
                )
        columns.append(
            ColumnMapping(
                csv_column=csv_col,
                excel_column=str(excel_col).upper(),
                number_format=item.get("number_format"),
                align=align,
            )
        )
    if not columns:
        raise ValueError("columns: 1件以上の列マッピングが必要です")
    return columns


def _parse_sort(raw: list[dict[str, Any]] | None) -> list[SortRule]:
    if not raw:
        return []
    rules: list[SortRule] = []
    for index, item in enumerate(raw):
        csv_col = _require(item, "csv", f"sort[{index}]")
        order = str(_require(item, "order", f"sort[{index}]")).lower()
        if order not in ("asc", "desc"):
            raise ValueError(f"sort[{index}].order は 'asc' または 'desc' を指定してください")
        rules.append(SortRule(csv_column=csv_col, order=order))
    return rules


def _parse_horizontal_borders(raw: list[dict[str, Any]] | None) -> list[HorizontalBorder]:
    if not raw:
        return []
    borders: list[HorizontalBorder] = []
    for index, item in enumerate(raw):
        row = _require(item, "row", f"borders.horizontal[{index}]")
        if row != "last_data_row":
            row = int(row)
        from_col = str(_require(item, "from_col", f"borders.horizontal[{index}]")).upper()
        to_col = str(_require(item, "to_col", f"borders.horizontal[{index}]")).upper()
        column_index_from_string(from_col)
        column_index_from_string(to_col)
        borders.append(
            HorizontalBorder(
                row=row,
                from_col=from_col,
                to_col=to_col,
                style=item.get("style", "thin"),
            )
        )
    return borders


def _column_span(columns: list[ColumnMapping]) -> tuple[str, str]:
    indices = [column_index_from_string(col.excel_column) for col in columns]
    return get_column_letter(min(indices)), get_column_letter(max(indices))


def _parse_border_region(item: dict[str, Any], context: str) -> BorderRegion:
    from_row = int(_require(item, "from_row", context))
    to_row = _require(item, "to_row", context)
    if to_row != "last_data_row":
        to_row = int(to_row)
    from_col = str(_require(item, "from_col", context)).upper()
    to_col = str(_require(item, "to_col", context)).upper()
    column_index_from_string(from_col)
    column_index_from_string(to_col)
    mode = str(_require(item, "mode", context)).lower()
    if mode not in BORDER_MODES:
        allowed = ", ".join(BORDER_MODES)
        raise ValueError(f"{context}.mode は {allowed} のいずれかを指定してください")
    row_step = int(item.get("row_step", 1))
    col_step = int(item.get("col_step", 1))
    if row_step < 1:
        raise ValueError(f"{context}.row_step は 1 以上を指定してください")
    if col_step < 1:
        raise ValueError(f"{context}.col_step は 1 以上を指定してください")
    return BorderRegion(
        from_row=from_row,
        to_row=to_row,
        from_col=from_col,
        to_col=to_col,
        mode=mode,
        style=item.get("style", "thin"),
        internal=bool(item.get("internal", True)),
        row_step=row_step,
        col_step=col_step,
    )


def _parse_border_regions(raw: list[dict[str, Any]] | None) -> list[BorderRegion]:
    if not raw:
        return []
    regions: list[BorderRegion] = []
    for index, item in enumerate(raw):
        regions.append(_parse_border_region(item, f"borders.regions[{index}]"))
    return regions


def _parse_data_table_border(
    raw: dict[str, Any] | None,
    columns: list[ColumnMapping],
) -> BorderRegion | None:
    if not raw:
        return None
    context = "borders.data_table"
    header_row = int(_require(raw, "header_row", context))
    mode = str(_require(raw, "mode", context)).lower()
    if mode not in BORDER_MODES:
        allowed = ", ".join(BORDER_MODES)
        raise ValueError(f"{context}.mode は {allowed} のいずれかを指定してください")
    from_col, to_col = _column_span(columns)
    to_row = raw.get("to_row", "last_data_row")
    if to_row != "last_data_row":
        to_row = int(to_row)
    row_step = int(raw.get("row_step", 1))
    col_step = int(raw.get("col_step", 1))
    if row_step < 1:
        raise ValueError(f"{context}.row_step は 1 以上を指定してください")
    if col_step < 1:
        raise ValueError(f"{context}.col_step は 1 以上を指定してください")
    if "from_col" in raw:
        from_col = str(raw["from_col"]).upper()
        column_index_from_string(from_col)
    if "to_col" in raw:
        to_col = str(raw["to_col"]).upper()
        column_index_from_string(to_col)
    return BorderRegion(
        from_row=header_row,
        to_row=to_row,
        from_col=from_col,
        to_col=to_col,
        mode=mode,
        style=raw.get("style", "thin"),
        internal=bool(raw.get("internal", True)),
        row_step=row_step,
        col_step=col_step,
    )


def _parse_vertical_borders(raw: list[dict[str, Any]] | None) -> list[VerticalBorder]:
    if not raw:
        return []
    borders: list[VerticalBorder] = []
    for index, item in enumerate(raw):
        col = str(_require(item, "col", f"borders.vertical[{index}]")).upper()
        column_index_from_string(col)
        from_row = int(_require(item, "from_row", f"borders.vertical[{index}]"))
        to_row = _require(item, "to_row", f"borders.vertical[{index}]")
        if to_row != "last_data_row":
            to_row = int(to_row)
        borders.append(
            VerticalBorder(
                col=col,
                from_row=from_row,
                to_row=to_row,
                style=item.get("style", "thin"),
            )
        )
    return borders


def load_config(config_path: str | Path) -> ReportConfig:
    path = resolve_app_path(config_path)
    if not path.exists():
        raise FileNotFoundError(f"設定ファイルが見つかりません: {path}")

    with path.open(encoding="utf-8") as f:
        raw = yaml.safe_load(f) or {}

    app_root = get_app_root()
    paths = raw.get("paths", {})
    data = raw.get("data", {})
    borders = raw.get("borders", {})
    output = raw.get("output", {})
    csv = raw.get("csv", {})

    template_path = resolve_app_path(_require(paths, "template", "paths"))
    csv_input_dir = resolve_app_path(_require(paths, "csv_input", "paths"))
    excel_output_dir = resolve_app_path(_require(paths, "excel_output", "paths"))

    columns = _parse_columns(_require(data, "columns", "data"))
    border_regions = _parse_border_regions(borders.get("regions"))
    data_table_region = _parse_data_table_border(borders.get("data_table"), columns)
    if data_table_region is not None:
        border_regions.append(data_table_region)

    return ReportConfig(
        name=raw.get("name", path.stem),
        config_path=path,
        base_dir=app_root,
        csv_input_dir=csv_input_dir,
        excel_output_dir=excel_output_dir,
        template_path=template_path,
        output=OutputSettings(
            base_name=output.get("base_name", raw.get("name", "report")),
            include_timestamp=bool(output.get("include_timestamp", True)),
            timestamp_format=output.get("timestamp_format", "%Y%m%d_%H%M%S"),
        ),
        csv=CsvSettings(
            encoding=csv.get("encoding", "utf-8-sig"),
            delimiter=csv.get("delimiter", ","),
            pattern=csv.get("pattern", "*.csv"),
        ),
        sheet_name=data.get("sheet_name"),
        start_row=int(data.get("start_row", 2)),
        columns=columns,
        sort_rules=_parse_sort(data.get("sort")),
        horizontal_borders=_parse_horizontal_borders(borders.get("horizontal")),
        vertical_borders=_parse_vertical_borders(borders.get("vertical")),
        border_regions=border_regions,
    )
