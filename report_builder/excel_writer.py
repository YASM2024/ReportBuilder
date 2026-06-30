"""CSV読み込み・ソート・Excel書き込み。"""

from __future__ import annotations

import csv
from datetime import datetime
from pathlib import Path
from typing import Any

import chardet
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter

from .config_loader import (
    BorderRegion,
    ColumnMapping,
    HorizontalBorder,
    ReportConfig,
    SortRule,
    VerticalBorder,
)


def detect_encoding(path: Path) -> str:
    raw = path.read_bytes()
    detected = chardet.detect(raw)
    encoding = detected.get("encoding") or "utf-8"
    if encoding.lower().replace("-", "") in ("cp932", "shiftjis", "mskanji"):
        return "cp932"
    return encoding


def read_csv(path: Path, encoding: str, delimiter: str) -> tuple[list[str], list[dict[str, str]]]:
    used_encoding = encoding
    if encoding == "auto":
        used_encoding = detect_encoding(path)

    with path.open(encoding=used_encoding, newline="") as f:
        reader = csv.DictReader(f, delimiter=delimiter)
        if reader.fieldnames is None:
            raise ValueError(f"CSVにヘッダ行がありません: {path}")
        headers = [h.strip() if h else "" for h in reader.fieldnames]
        rows: list[dict[str, str]] = []
        for row in reader:
            normalized = {
                (key.strip() if key else ""): (value if value is not None else "")
                for key, value in row.items()
            }
            rows.append(normalized)
    return headers, rows


def _resolve_csv_value(row: dict[str, str], headers: list[str], column: str | int) -> str:
    if isinstance(column, int):
        if column < 0 or column >= len(headers):
            raise IndexError(f"CSV列インデックス {column} が範囲外です")
        return row.get(headers[column], "")
    if column not in row and column not in headers:
        raise KeyError(f"CSV列 '{column}' が見つかりません。利用可能: {headers}")
    return row.get(column, "")


def _sort_key(value: str) -> tuple[int, Any]:
    stripped = value.strip()
    if stripped == "":
        return (2, "")
    try:
        return (0, float(stripped.replace(",", "")))
    except ValueError:
        return (1, stripped)


def sort_rows(
    rows: list[dict[str, str]],
    headers: list[str],
    rules: list[SortRule],
) -> list[dict[str, str]]:
    if not rules:
        return rows

    def key_func(row: dict[str, str]) -> tuple:
        keys = []
        for rule in rules:
            value = _resolve_csv_value(row, headers, rule.csv_column)
            keys.append(_sort_key(value))
            if rule.order == "desc":
                keys[-1] = (keys[-1][0], _invert_sortable(keys[-1][1]))
        return tuple(keys)

    return sorted(rows, key=key_func)


def _invert_sortable(value: Any) -> Any:
    if isinstance(value, (int, float)):
        return -value
    return value


def _side(style_name: str) -> Side:
    return Side(style=style_name, color="000000")


def _resolve_row(row: int | str, last_data_row: int) -> int:
    return last_data_row if row == "last_data_row" else int(row)


def _apply_border_region(ws, region: BorderRegion, last_data_row: int) -> None:
    from_row = region.from_row
    to_row = _resolve_row(region.to_row, last_data_row)
    if to_row < from_row:
        return

    from_col_idx = column_index_from_string(region.from_col)
    to_col_idx = column_index_from_string(region.to_col)
    side = _side(region.style)

    if region.mode == "outline":
        draw_horizontal = True
        draw_vertical = True
        horizontal_internal = False
        vertical_internal = False
    elif region.mode == "horizontal":
        draw_horizontal = True
        draw_vertical = False
        horizontal_internal = region.internal
        vertical_internal = False
    elif region.mode == "vertical":
        draw_horizontal = False
        draw_vertical = True
        horizontal_internal = False
        vertical_internal = region.internal
    else:  # grid
        draw_horizontal = True
        draw_vertical = True
        horizontal_internal = region.internal
        vertical_internal = region.internal

    for row_idx in range(from_row, to_row + 1):
        for col_idx in range(from_col_idx, to_col_idx + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            existing = cell.border
            top = existing.top
            bottom = existing.bottom
            left = existing.left
            right = existing.right

            if draw_horizontal:
                if row_idx == from_row:
                    top = side
                if row_idx == to_row:
                    bottom = side
                if horizontal_internal and row_idx < to_row:
                    if (row_idx - from_row) % region.row_step == 0:
                        bottom = side

            if draw_vertical:
                if col_idx == from_col_idx:
                    left = side
                if col_idx == to_col_idx:
                    right = side
                if vertical_internal and col_idx < to_col_idx:
                    if (col_idx - from_col_idx) % region.col_step == 0:
                        right = side

            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _apply_horizontal_border(ws, border: HorizontalBorder, last_data_row: int) -> None:
    row = _resolve_row(border.row, last_data_row)
    from_idx = column_index_from_string(border.from_col)
    to_idx = column_index_from_string(border.to_col)
    side = _side(border.style)
    for col_idx in range(from_idx, to_idx + 1):
        cell = ws.cell(row=row, column=col_idx)
        existing = cell.border
        cell.border = Border(
            left=existing.left,
            right=existing.right,
            top=side,
            bottom=side,
        )


def _apply_vertical_border(ws, border: VerticalBorder, last_data_row: int) -> None:
    col_idx = column_index_from_string(border.col)
    end_row = _resolve_row(border.to_row, last_data_row)
    side = _side(border.style)
    for row_idx in range(border.from_row, end_row + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        existing = cell.border
        cell.border = Border(
            left=side,
            right=side,
            top=existing.top,
            bottom=existing.bottom,
        )


def _apply_cell_alignment(cell, horizontal: str) -> None:
    existing = cell.alignment
    cell.alignment = Alignment(
        horizontal=horizontal,
        vertical=existing.vertical,
        text_rotation=existing.text_rotation,
        wrap_text=existing.wrap_text,
        shrink_to_fit=existing.shrink_to_fit,
        indent=existing.indent,
    )


def _coerce_cell_value(raw: str) -> str | int | float:
    stripped = raw.strip()
    if stripped == "":
        return ""
    try:
        if "." in stripped or "e" in stripped.lower():
            return float(stripped.replace(",", ""))
        return int(stripped.replace(",", ""))
    except ValueError:
        return raw


def write_report(
    config: ReportConfig,
    csv_path: Path,
    output_path: Path | None = None,
) -> Path:
    headers, rows = read_csv(csv_path, config.csv.encoding, config.csv.delimiter)
    sorted_rows = sort_rows(rows, headers, config.sort_rules)

    if not config.template_path.exists():
        raise FileNotFoundError(f"Excelひな形が見つかりません: {config.template_path}")

    workbook = load_workbook(config.template_path)
    if config.sheet_name:
        if config.sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"シート '{config.sheet_name}' がひな形にありません: {workbook.sheetnames}"
            )
        worksheet = workbook[config.sheet_name]
    else:
        worksheet = workbook.active

    data_row_count = len(sorted_rows)
    last_data_row = config.start_row + data_row_count - 1 if data_row_count else config.start_row - 1

    for offset, row in enumerate(sorted_rows):
        excel_row = config.start_row + offset
        for mapping in config.columns:
            raw_value = _resolve_csv_value(row, headers, mapping.csv_column)
            col_idx = column_index_from_string(mapping.excel_column)
            cell = worksheet.cell(row=excel_row, column=col_idx)
            cell.value = _coerce_cell_value(raw_value)
            if mapping.number_format:
                cell.number_format = mapping.number_format
            if mapping.align:
                _apply_cell_alignment(cell, mapping.align)

    for region in config.border_regions:
        _apply_border_region(worksheet, region, last_data_row)

    for border in config.horizontal_borders:
        _apply_horizontal_border(worksheet, border, last_data_row)

    for border in config.vertical_borders:
        _apply_vertical_border(worksheet, border, last_data_row)

    if output_path is None:
        output_path = build_output_path(config, csv_path)

    config.excel_output_dir.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return output_path


def build_output_path(config: ReportConfig, csv_path: Path) -> Path:
    if config.output.include_timestamp:
        timestamp = datetime.now().strftime(config.output.timestamp_format)
        filename = f"{config.output.base_name}_{timestamp}.xlsx"
    else:
        filename = f"{config.output.base_name}.xlsx"
    return config.excel_output_dir / filename


def find_csv_files(config: ReportConfig) -> list[Path]:
    if not config.csv_input_dir.exists():
        raise FileNotFoundError(f"CSV入力フォルダが見つかりません: {config.csv_input_dir}")
    files = sorted(config.csv_input_dir.glob(config.csv.pattern))
    if not files:
        raise FileNotFoundError(
            f"CSVファイルが見つかりません: {config.csv_input_dir / config.csv.pattern}"
        )
    return files
